from openpyxl import load_workbook

file1 = load_workbook(r'../实体关系表.xlsx')
ac_sheet1 = file1['Sheet1']

f1 = open(r'train_labeled.txt', 'r', encoding='utf-8')
f2 = open(r'test_labeled.txt', 'r', encoding='utf-8')
def get_sen(f):
    sen = []
    while True:
        con = f.readline()
        if con == '':
            break
        line = con.strip()
        sen.append(line)
    return list(set(sen))

all_sen = []
sen1 = get_sen(f1)
sen2 = get_sen(f2)

for i in range(len(sen1)):
    all_sen.append(sen1[i])
for i in range(len(sen2)):
    all_sen.append(sen2[i])

# print(all_sen)
for i in range(len(all_sen)):
    ac_sheet1.cell(row=i+2, column=1, value=all_sen[i].strip().split()[0])
    ac_sheet1.cell(row=i+2, column=2, value=all_sen[i].strip().split()[1])
    ac_sheet1.cell(row=i+2, column=3, value=all_sen[i].strip().split()[2])
    ac_sheet1.cell(row=i+2, column=4, value=all_sen[i].strip().split()[3])

file1.save('../实体关系表.xlsx')
