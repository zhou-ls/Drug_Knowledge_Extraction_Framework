# -*- coding: utf-8 -*-
# @Time : 2020/11/5 10:05
# @Author : zls
# @File : get_train.py
from openpyxl import load_workbook

file = load_workbook('实体关系表.xlsx')
ac_sheet = file['Sheet1']

num1 = 0
num2 = 0
num3 = 0
num4 = 0
f = open(r'all.txt', 'w', encoding='utf-8')
for i in range(2, 6040):
    en1 = ac_sheet.cell(row=i, column=1).value
    en2 = ac_sheet.cell(row=i, column=2).value
    ex = ac_sheet.cell(row=i, column=3).value
    sen = ac_sheet.cell(row=i, column=4).value
    if ex == '药效降低' and num1 < 420:
        num1 += 1
        f.write(en1 + '\t' + en2 + '\t' + ex + '\t' + sen + '\n')
    elif ex == '药效提高' and num2 < 420:
        num2 += 1
        f.write(en1 + '\t' + en2 + '\t' + ex + '\t' + sen + '\n')
    elif ex == '药效提高，不良反应增加' and num3 < 420:
        num3 += 1
        f.write(en1 + '\t' + en2 + '\t' + ex + '\t' + sen + '\n')
    elif ex == '不良反应增加' and num4 < 420:
        num4 += 1
        f.write(en1 + '\t' + en2 + '\t' + ex + '\t' + sen + '\n')
    elif ex == '药效降低，不良反应增加':
        f.write(en1 + '\t' + en2 + '\t' + ex + '\t' + sen + '\n')
